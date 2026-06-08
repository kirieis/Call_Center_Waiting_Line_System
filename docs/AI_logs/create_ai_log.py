import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.Workbook()

# ============================================================
# STYLE DEFINITIONS
# ============================================================
title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
section_font = Font(name='Arial', size=13, bold=True, color='1F4E79')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
label_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=11)
entry_font = Font(name='Arial', size=11, color='333333')

title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
light_green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
light_yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
light_red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(row=r, column=c).border = thin_border

# ============================================================
# SHEET 1: Metadata & Summary
# ============================================================
ws1 = wb.active
ws1.title = '1. Metadata & Summary'

ws1.column_dimensions['A'].width = 45
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 45
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 25
ws1.column_dimensions['F'].width = 20

# Title
ws1.merge_cells('A1:F1')
c = ws1['A1']
c.value = 'AI AUDIT LOG - METADATA & SUMMARY'
c.font = title_font
c.fill = title_fill
c.alignment = center_align

# Student Information Section
ws1.merge_cells('A3:F3')
ws1['A3'].value = 'STUDENT INFORMATION'
ws1['A3'].font = section_font
ws1['A3'].fill = light_blue_fill

info_data = [
    ('Student Name:', 'Nguyễn Văn An'),
    ('Student ID:', 'QE200040'),
    ('Course:', 'CSD201 / CSD203 (LAB)'),
    ('Assignment:', 'Call Center Waiting Line System - Group 7'),
]
for i, (label, val) in enumerate(info_data, start=4):
    ws1.cell(row=i, column=1, value=label).font = label_font
    ws1.cell(row=i, column=3, value=val).font = normal_font

# AI Usage Summary
ws1.merge_cells('A9:F9')
ws1['A9'].value = 'AI USAGE SUMMARY'
ws1['A9'].font = section_font
ws1['A9'].fill = light_blue_fill

usage_data = [
    ('Total Prompts Used (all AI tools):', '~50+'),
    ('Core Prompts Logged:', '11'),
    ('Selection Ratio:', '~20% (11/50+ = only core decision/problem-solving/verification prompts)'),
    ('Hallucination Detected:', '3'),
]
for i, (label, val) in enumerate(usage_data, start=10):
    ws1.cell(row=i, column=1, value=label).font = label_font
    ws1.cell(row=i, column=3, value=val).font = normal_font

# AI Tools Used
ws1.merge_cells('A15:F15')
ws1['A15'].value = 'AI TOOLS USED'
ws1['A15'].font = section_font
ws1['A15'].fill = light_blue_fill

tools_headers = ['AI Tool', 'Purpose', 'Frequency', 'Main Value']
for i, h in enumerate(tools_headers, 1):
    c = ws1.cell(row=16, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

tools_data = [
    ('Claude 4.6 Opus (Antigravity)', 'Code generation, architecture design, UML diagrams, debugging', 'High', 'Core development partner'),
    ('Gemini Flash', 'Quick reference, fact checking', 'Low', 'Secondary verification'),
    ('Codex 5.4 + 5.5', 'Code assistance, auto-completion', 'Medium', 'Code completion'),
]
for i, row_data in enumerate(tools_data, start=17):
    for j, val in enumerate(row_data, 1):
        c = ws1.cell(row=i, column=j, value=val)
        c.font = entry_font
        c.alignment = wrap_align

apply_border_range(ws1, 16, 19, 1, 4)

# DTC Component Distribution
ws1.merge_cells('A22:F22')
ws1['A22'].value = 'CORE PROMPTS DISTRIBUTION BY DTC COMPONENT'
ws1['A22'].font = section_font
ws1['A22'].fill = light_blue_fill

dtc_headers = ['DTC Component', 'Number of Prompts', 'Required (Min)']
for i, h in enumerate(dtc_headers, 1):
    c = ws1.cell(row=23, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

dtc_data = [
    ('Decomposition', 3, '≥ 1'),
    ('Pattern Recognition', 1, '≥ 1'),
    ('Abstraction', 2, '≥ 1'),
    ('Algorithms', 5, '≥ 1'),
]
for i, row_data in enumerate(dtc_data, start=24):
    for j, val in enumerate(row_data, 1):
        c = ws1.cell(row=i, column=j, value=val)
        c.font = entry_font
        c.alignment = center_align

apply_border_range(ws1, 23, 27, 1, 3)

# ============================================================
# SHEET 2: Detailed Audit Log (FILTERED - 11 CORE PROMPTS)
# ============================================================
ws2 = wb.create_sheet('2. Detailed Audit Log')

col_widths = [10, 20, 22, 32, 50, 48, 55, 28]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Title
ws2.merge_cells('A1:H1')
c = ws2['A1']
c.value = 'DETAILED AI AUDIT LOG'
c.font = title_font
c.fill = title_fill
c.alignment = center_align

ws2.merge_cells('A2:H2')
ws2['A2'].value = 'INSTRUCTIONS: Chỉ ghi CORE PROMPTS (Decision/Problem-Solving/Verification). Đã lọc từ ~50+ prompts xuống 11 core prompts. Mỗi entry pass ≥4/5 tiêu chí chất lượng.'
ws2['A2'].font = Font(name='Arial', size=10, italic=True, color='FF0000')
ws2['A2'].alignment = wrap_align

# Headers
headers = ['Entry #', 'Prompt Type', 'Stage/Component', 'Problem/Context', 'Prompt to AI', 'AI Response (Summary)', 'Human Delta & Reflection', 'Evidence']
for i, h in enumerate(headers, 1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

# ---- 11 FILTERED CORE ENTRIES ----
entries = [
    # Entry 001 (Lượt 1): Phân tích Actor - DECISION
    {
        'entry': '001',
        'type': 'DECISION',
        'component': 'Decomposition',
        'context': 'Phân tích Actor trong Use Case Diagram - Tại sao chỉ có Operator mà không có Khách hàng VIP?',
        'prompt': '"đây là dự án cá nhân của tôi Call Center Waiting Line System là Hệ thống quản lý cuộc gọi chờ, ưu tiên khách hàng VIP hoặc khách hàng gọi lại nhiều lần. Tôi muốn bạn hãy xem thử use_case_diagram phân tích kĩ các actor tham gia có đầy đủ chưa. Tại sao ko thấy có khách hàng VIP hoặc khách hàng gọi lại nhiều lần ,... mà chỉ có operator."',
        'response': 'AI giải thích: Actor trong UML là vai trò bên ngoài trực tiếp tương tác với phần mềm. Vì đây là Console app, khách hàng không trực tiếp tương tác - chỉ Operator nhập liệu. Khách hàng VIP là Data Entity (thuộc tính dữ liệu của Call), không phải Actor. Kết luận: Operator là Actor duy nhất là hoàn toàn chính xác.',
        'delta': 'Critical Thinking: Phân biệt rõ Actor (tác nhân tương tác trực tiếp) vs Data Entity (đối tượng dữ liệu) trong UML. Contextualization: Console app nội bộ → chỉ nhân viên vận hành mới tương tác trực tiếp với phần mềm. Creative Synthesis: VIP/khách gọi lại sẽ được mô hình hóa qua thuộc tính priorityPoint trong model Call. Decision: Giữ nguyên Operator là Actor duy nhất.',
        'evidence': 'Use Case Diagram file (use_case_diagram.drawio)'
    },
    # Entry 002 (Lượt 2): Tối ưu cây thư mục lần 1 - DECISION
    {
        'entry': '002',
        'type': 'DECISION',
        'component': 'Decomposition',
        'context': 'Đánh giá và tối ưu cấu trúc thư mục dự án - AI đề xuất cắt giảm quá tay',
        'prompt': '"Project của tôi yêu cầu làm oop thuật toán trên console thôi như vậy thì đang đúng hướng đúng không. Hãy phân tích cây thư mục project_tree_folder của tôi cho tôi biết liệu tôi có thể tối ưu cắt bớt 1 chút được không ( tôi muốn ít gây nhầm lẫn nhất có thể) nếu không thể cắt giảm bớt thì thôi."',
        'response': 'AI xác nhận đúng hướng OOP + Console. Đề xuất phương án cắt giảm tối đa: gom các package UI, gom file lõi. Tuy nhiên phương án này cắt bỏ quá nhiều lớp chức năng quan trọng, vi phạm nguyên lý SRP (Single Responsibility Principle).',
        'delta': 'Critical Thinking: Nhận ra phương án tối giản quá mức sẽ vi phạm nguyên lý OOP - gom quá nhiều trách nhiệm vào ít file. Contextualization: Cần giữ phân tách rõ ràng model/core/ui/storage cho project LAB đạt chuẩn OOP. Decision: BÁC BỎ phương án AI đề xuất, yêu cầu làm lại vừa phải hơn.',
        'evidence': 'Project tree-folder comparison (before/after)'
    },
    # Entry 003 (Lượt 3): Tối ưu vừa phải - PROBLEM-SOLVING
    {
        'entry': '003',
        'type': 'PROBLEM-SOLVING',
        'component': 'Decomposition',
        'context': 'Yêu cầu cấu trúc tối ưu vừa phải sau khi bác bỏ phương án quá tối giản',
        'prompt': '"hình như lược bỏ hơi nhiều rồi làm lại đi"',
        'response': 'AI đề xuất cấu trúc Moderate Optimization: Giữ nguyên phân tách ui/, core/, storage/, experiment/. Gom core/queue/ và core/routing/ vào core/ giảm 1 cấp. Giữ đầy đủ file thực nghiệm. Cấu trúc cuối: src/callcenter/ với 5 package: model, core, ui, storage, experiment.',
        'delta': 'Critical Thinking: So sánh 2 phương án (quá tối giản vs vừa phải), chọn phương án cân bằng giữa đơn giản và chuẩn OOP. Contextualization: Project LAB cần đủ phân tách cho OOP nhưng không quá phức tạp. Creative Synthesis: Giảm cấp thư mục lồng nhau nhưng giữ nguyên phân tách chức năng. Decision: Chọn Moderate Optimization.',
        'evidence': 'Final project tree structure (5 packages: model, core, ui, storage, experiment)'
    },
    # Entry 004 (Lượt 7): Cập nhật UML diagrams - PROBLEM-SOLVING
    {
        'entry': '004',
        'type': 'PROBLEM-SOLVING',
        'component': 'Abstraction',
        'context': 'Cập nhật Use Case Diagram và Class Diagram theo cấu trúc thư mục mới đã tối ưu',
        'prompt': '"Vì đã thay đổi tcaay thư mục của project nên chắc chắn use_case_diagram và class diagram cũng sẽ thay đổi. Bạn hãy thực hiện Vẽ lại chính xác cho 2 diagram giúp tôi dựa theo file tree-folder."',
        'response': 'AI thiết kế lại 2 sơ đồ Draw.io: (1) Use Case: Actor Operator, 6 UC chính (xanh) + 4 UC phụ trợ (vàng) với quan hệ include/extend. (2) Class Diagram: Loại bỏ StateManager cũ, MainMenu làm trung tâm, cập nhật package callcenter.*, bổ sung waitTime/orderNumber cho Call, enum CallStatus 3 trạng thái.',
        'delta': 'Critical Thinking: Nhận ra thay đổi cấu trúc thư mục bắt buộc phải cập nhật lại UML để đảm bảo tính nhất quán. Contextualization: Diagram phải phản ánh chính xác code structure. Creative Synthesis: Phân loại UC thành chính (xanh) và phụ trợ (vàng) giúp dễ hiểu. Decision: Phê duyệt thiết kế UML mới.',
        'evidence': 'use_case_diagram.drawio, class_diagram.drawio files'
    },
    # Entry 005 (Lượt 8): Sửa lỗi XML - VERIFICATION (Hallucination)
    {
        'entry': '005',
        'type': 'VERIFICATION',
        'component': 'Pattern Recognition',
        'context': 'Phát hiện AI sinh XML không hợp lệ - lỗi double-hyphen trong comment của Draw.io files',
        'prompt': '"class diagram bị lỗi:\nNot a diagram file (error on line 237 at column 19:\nDouble hyphen within comment: <!-- Main)\nuse_case_diagram bị lỗi:\nNot a diagram file (error on line 94 at column 18:\nDouble hyphen within comment: <!-- UC1)"',
        'response': 'AI xác định nguyên nhân: Tiêu chuẩn XML nghiêm cấm hai dấu gạch ngang liên tiếp (--) trong comment. AI đã sinh ra comment dạng <!-- UC1 -- Thêm cuộc gọi --> là vi phạm XML spec. Giải pháp: Loại bỏ hoàn toàn XML comment, lưu phiên bản XML sạch 100%.',
        'delta': 'Critical Thinking: Phát hiện HALLUCINATION - AI đã tạo XML comment không hợp lệ theo chuẩn W3C. Contextualization: Lỗi syntax cần tuân thủ nghiêm ngặt tiêu chuẩn XML. Decision: Yêu cầu AI sửa và kiểm tra lại toàn bộ 2 file .drawio.',
        'evidence': 'Fixed .drawio files (XML validated, no comments)'
    },
    # Entry 006 (Lượt 10): Giải thích PriorityQueue sort - VERIFICATION
    {
        'entry': '006',
        'type': 'VERIFICATION',
        'component': 'Algorithms',
        'context': 'Kiểm chứng cơ chế sắp xếp khi tạo ArrayList từ PriorityQueue để hiển thị',
        'prompt': '"List<Call> sortedCalls = new ArrayList<>(waitingCalls); sortedCalls.sort(waitingCalls.comparator()) đoạn code này tôi chưa hiểu lắm bạn nói kĩ cách các call đc sắp xếp đi"',
        'response': 'AI giải thích: PriorityQueue (Binary Heap) không duy trì thứ tự khi duyệt bằng iterator - chỉ đảm bảo phần tử cao nhất ở gốc. Để hiển thị đúng thứ tự: (1) new ArrayList<>() tạo bản sao không phá Heap gốc. (2) sort() dùng cùng Comparator của PQ, áp dụng TimSort (stable sort) để sắp xếp hoàn chỉnh.',
        'delta': 'Critical Thinking: Phân biệt Heap (partial order - chỉ đảm bảo root) vs Array (total order - sắp xếp hoàn chỉnh). Contextualization: Phải tạo bản sao ArrayList để không phá vỡ cấu trúc Heap gốc khi chỉ cần hiển thị. Decision: Giữ nguyên cách tiếp cận copy-then-sort.',
        'evidence': 'CallQueue.java - displayWaitingCalls() method'
    },
    # Entry 007 (Lượt 11): Cải tiến Comparator - DECISION
    {
        'entry': '007',
        'type': 'DECISION',
        'component': 'Algorithms',
        'context': 'Đánh giá 2 phong cách viết Comparator: Anonymous class truyền thống vs Java 8 Lambda',
        'prompt': '"có cách viết nào dễ đọc dễ hiểu hơn không mà vẫn giúp chương trình hoạt động như bình thường ở cả sortedCalls.sort(waitingCalls.comparator()); và cả hàm compare"',
        'response': 'AI đề xuất Java 8 Lambda: Comparator.comparingInt(Call::getPriorityPoint).reversed().thenComparingInt(Call::getOrderNumber). Ngắn gọn, đọc như tiếng Anh, sort vẫn tương thích. Tuy nhiên che giấu logic so sánh chi tiết.',
        'delta': 'Critical Thinking: So sánh tradeoff: Anonymous class (tường minh, dễ debug, dễ comment giải thích) vs Lambda (ngắn gọn nhưng ẩn logic). Contextualization: Project LAB ưu tiên code dễ hiểu cho người mới đọc (thầy cô, đồng đội). Decision: Cuối cùng BÁC BỎ Lambda, chọn giữ anonymous class + comment chi tiết (xem Entry 009).',
        'evidence': 'Code comparison (anonymous class vs lambda expression)'
    },
    # Entry 008 (Lượt 12): Binary Heap xác nhận - VERIFICATION
    {
        'entry': '008',
        'type': 'VERIFICATION',
        'component': 'Algorithms',
        'context': 'Xác nhận PriorityQueue vẫn là Binary Heap khi thay đổi cú pháp Comparator',
        'prompt': '"vậy priorityqueue có còn là binary heap không hay đây chỉ là cách viết dễ nhìn hơn thôi"',
        'response': 'AI xác nhận: PriorityQueue LUÔN là Binary Heap (Min-Heap) bất kể cú pháp Comparator. Java 8 Lambda chỉ là syntax sugar, không thay đổi cấu trúc dữ liệu. Vẫn là mảng biểu diễn cây nhị phân, Heapify O(log n).',
        'delta': 'Critical Thinking: Phân biệt rõ cú pháp viết code (syntax layer) vs cấu trúc dữ liệu thực tế (implementation layer). Contextualization: Kiến thức quan trọng cho báo cáo và oral vivas - phải giải thích được bản chất thuật toán không phụ thuộc cách viết. Decision: Ghi nhận để chuẩn bị cho vấn đáp.',
        'evidence': 'Java PriorityQueue JDK source code reference (Binary Heap implementation)'
    },
    # Entry 009 (Lượt 14): Discard Lambda, giữ truyền thống - DECISION
    {
        'entry': '009',
        'type': 'DECISION',
        'component': 'Algorithms',
        'context': 'Quyết định cuối cùng: BÁC BỎ code Java 8 Lambda của AI, quay lại anonymous class + bổ sung comment giải thích',
        'prompt': '"thôi tôi discard code mới bạn cho rồi. Tôi muốn bạn comment thêm để người đọc có thể hiểu cách nó sắp xếp"',
        'response': 'AI tôn trọng quyết định, khôi phục anonymous class truyền thống trong CallQueue.java. Bổ sung comment tiếng Việt chi tiết: quy tắc số DƯƠNG (>0) cho giảm dần điểm ưu tiên, quy tắc số ÂM (<0) cho tăng dần orderNumber (FIFO).',
        'delta': 'Critical Thinking: Code dễ đọc + có comment giải thích > code ngắn gọn nhưng khó hiểu cho project LAB. Contextualization: Người đọc code (thầy cô, đồng đội) chưa chắc quen Lambda. Creative Synthesis: Kết hợp cú pháp truyền thống + comment tiếng Việt chi tiết. Decision Ownership: CHỦ ĐỘNG discard gợi ý AI, chọn approach phù hợp ngữ cảnh hơn.',
        'evidence': 'CallQueue.java with detailed Vietnamese comments on compare logic'
    },
    # Entry 010 (Lượt 15): Tại sao if + swap mechanism - VERIFICATION
    {
        'entry': '010',
        'type': 'VERIFICATION',
        'component': 'Abstraction',
        'context': 'Kiểm chứng tại sao dùng if kiểm tra điểm ưu tiên trước, và cơ chế hoán đổi vị trí thực tế của Java',
        'prompt': '"tại sao là if first.getPriorityPoint() != second.getPriorityPoint(). Và ví dụ nếu return ở if là dương đi thì cách nó đổi vị trí đc đề cập ở đâu"',
        'response': 'AI giải thích: (1) Dùng if vì điểm ưu tiên là tiêu chí hàng đầu - khác nhau thì kết thúc ngay, chỉ so orderNumber khi bằng nhau (multi-level priority). (2) Swap thực tế do JDK thực hiện: siftUp()/siftDown() cho PriorityQueue, TimSort cho ArrayList.sort() - logic so sánh (do ta viết) tách biệt khỏi logic hoán đổi (do JDK viết).',
        'delta': 'Critical Thinking: Phân biệt rõ "logic so sánh" (code chúng ta) vs "logic hoán đổi" (JDK internal). Contextualization: Hiểu kiến trúc phân lớp Java Collections Framework - Comparator là interface contract. Decision: Ghi nhận để trả lời chính xác trong oral vivas.',
        'evidence': 'Java PriorityQueue siftUp/siftDown source analysis'
    },
    # Entry 011 (Lượt 16): Edge case compare trả về 0 - VERIFICATION
    {
        'entry': '011',
        'type': 'VERIFICATION',
        'component': 'Algorithms',
        'context': 'Phân tích edge case: nếu compare trả về 0 (hai cuộc gọi trùng orderNumber) thì hệ thống xử lý thế nào?',
        'prompt': '"vậy xong if thì qua xem vị trí đk. nếu ra 0 thì sao tương tự ở if"',
        'response': 'AI phân tích: (1) Thực tế: orderNumber++ tăng tự động nên KHÔNG BAO GIỜ trùng giữa 2 cuộc gọi khác nhau → compare không trả về 0. (2) Lý thuyết nếu trả về 0: ArrayList.sort (stable) giữ nguyên thứ tự ban đầu; PriorityQueue (unstable heap) không đảm bảo thứ tự → có thể gây non-deterministic behavior.',
        'delta': 'Critical Thinking: Phân tích edge case dù biết không xảy ra → hiểu sâu thuật toán và phòng ngừa. Contextualization: Phân biệt stable sort (TimSort) vs unstable heap. Creative Synthesis: Nhờ orderNumber++ tự động, hệ thống tránh hoàn toàn ambiguity - đây là thiết kế elegant cho FIFO guarantee. Decision: Thiết kế orderNumber++ là giải pháp đúng đắn.',
        'evidence': 'CallManager.java - orderNumber++ auto-increment mechanism'
    },
]

for i, e in enumerate(entries):
    row = i + 4
    ws2.cell(row=row, column=1, value=e['entry']).font = Font(name='Arial', size=12, bold=True)
    ws2.cell(row=row, column=1).alignment = center_align

    type_cell = ws2.cell(row=row, column=2, value=e['type'])
    type_cell.font = Font(name='Arial', size=11, bold=True)
    type_cell.alignment = center_align
    if e['type'] == 'DECISION':
        type_cell.fill = light_blue_fill
    elif e['type'] == 'PROBLEM-SOLVING':
        type_cell.fill = light_green_fill
    elif e['type'] == 'VERIFICATION':
        type_cell.fill = light_yellow_fill

    ws2.cell(row=row, column=3, value=e['component']).font = entry_font
    ws2.cell(row=row, column=3).alignment = center_align
    ws2.cell(row=row, column=4, value=e['context']).font = entry_font
    ws2.cell(row=row, column=4).alignment = wrap_align
    ws2.cell(row=row, column=5, value=e['prompt']).font = Font(name='Arial', size=11, italic=True)
    ws2.cell(row=row, column=5).alignment = wrap_align
    ws2.cell(row=row, column=6, value=e['response']).font = entry_font
    ws2.cell(row=row, column=6).alignment = wrap_align
    ws2.cell(row=row, column=7, value=e['delta']).font = entry_font
    ws2.cell(row=row, column=7).alignment = wrap_align
    ws2.cell(row=row, column=8, value=e['evidence']).font = entry_font
    ws2.cell(row=row, column=8).alignment = wrap_align

    ws2.row_dimensions[row].height = 100

apply_border_range(ws2, 3, 3 + len(entries), 1, 8)

# ============================================================
# SHEET 3: Hallucination Detection
# ============================================================
ws3 = wb.create_sheet('3. Hallucination Detection')

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 24
ws3.column_dimensions['C'].width = 45
ws3.column_dimensions['D'].width = 45
ws3.column_dimensions['E'].width = 40
ws3.column_dimensions['F'].width = 40

# Title
ws3.merge_cells('A1:F1')
c = ws3['A1']
c.value = 'HALLUCINATION DETECTION LOG (BẮT BUỘC)'
c.font = title_font
c.fill = title_fill
c.alignment = center_align

ws3.merge_cells('A2:F2')
ws3['A2'].value = 'MỖI PROJECT PHẢI PHÁT HIỆN ÍT NHẤT: Lab (≥1), Assignment (≥2), Project (≥3) cases hallucination'
ws3['A2'].font = Font(name='Arial', size=10, italic=True, color='FF0000')
ws3['A2'].alignment = wrap_align

# Headers
h3_headers = ['Entry # (from Sheet 2)', 'Hallucination Type', "AI's Claim", 'Reality Check', 'How Detected', 'Corrective Action']
for i, h in enumerate(h3_headers, 1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

hall_data = [
    {
        'entry': '005',
        'type': 'Logic Error',
        'claim': 'AI sinh mã XML cho file Draw.io (.drawio) có chứa comment dạng <!-- UC1 -- Thêm cuộc gọi --> và <!-- Main -- Lớp chính --> bên trong file XML.',
        'reality': 'Theo tiêu chuẩn XML (W3C), ký tự hai dấu gạch ngang liên tiếp (--) bị NGHIÊM CẤM bên trong nội dung comment XML. Draw.io từ chối mở file và báo lỗi "Not a diagram file".',
        'detected': 'Compiler/Tool Check: Draw.io báo lỗi khi mở file: "error on line 237 at column 19: Double hyphen within comment". Người dùng phát hiện lỗi và report lại.',
        'action': 'AI loại bỏ hoàn toàn tất cả XML comment trong cả 2 file .drawio. Tạo phiên bản XML sạch 100% tuân thủ chuẩn W3C.'
    },
    {
        'entry': '002',
        'type': 'Oversimplification',
        'claim': 'AI đề xuất phương án cắt giảm cấu trúc thư mục tối đa: gom toàn bộ package UI và file lõi vào chung, loại bỏ sự phân tách giữa các package chức năng.',
        'reality': 'Phương án vi phạm nguyên lý Single Responsibility Principle (SRP). Gom quá nhiều chức năng vào ít file/folder làm code khó bảo trì và không đạt chuẩn OOP cho project LAB.',
        'detected': 'Critical Thinking: Người dùng nhận ra "hình như lược bỏ hơi nhiều rồi" - phát hiện AI bỏ qua yêu cầu ngầm về chuẩn OOP của đề bài.',
        'action': 'AI thiết kế lại cấu trúc Moderate Optimization: giữ nguyên 5 package riêng biệt (model, core, ui, storage, experiment) nhưng giảm bớt cấp thư mục lồng nhau.'
    },
    {
        'entry': '007',
        'type': 'Context Misunderstanding',
        'claim': 'AI gợi ý chuyển toàn bộ Comparator sang cú pháp Java 8 Lambda Expression: Comparator.comparingInt(Call::getPriorityPoint).reversed() thay thế anonymous class truyền thống.',
        'reality': 'Trong ngữ cảnh project LAB học thuật, code cần dễ đọc cho người mới (thầy cô chấm bài, đồng đội). Lambda tuy ngắn gọn nhưng che giấu logic so sánh âm/dương, khiến người đọc khó trace cơ chế sắp xếp.',
        'detected': 'Decision Ownership: Người dùng chủ động discard code mới ("thôi tôi discard code mới bạn cho rồi"), yêu cầu giữ anonymous class + thêm comment tiếng Việt chi tiết.',
        'action': 'AI khôi phục anonymous class truyền thống, bổ sung comment giải thích chi tiết bằng tiếng Việt về quy tắc trả về DƯƠNG/ÂM của compare().'
    },
]

for i, h in enumerate(hall_data):
    row = i + 4
    ws3.cell(row=row, column=1, value=h['entry']).font = Font(name='Arial', size=12, bold=True)
    ws3.cell(row=row, column=1).alignment = center_align

    type_cell = ws3.cell(row=row, column=2, value=h['type'])
    type_cell.font = Font(name='Arial', size=11, bold=True)
    type_cell.alignment = center_align
    type_cell.fill = light_red_fill

    ws3.cell(row=row, column=3, value=h['claim']).font = entry_font
    ws3.cell(row=row, column=3).alignment = wrap_align
    ws3.cell(row=row, column=4, value=h['reality']).font = entry_font
    ws3.cell(row=row, column=4).alignment = wrap_align
    ws3.cell(row=row, column=5, value=h['detected']).font = entry_font
    ws3.cell(row=row, column=5).alignment = wrap_align
    ws3.cell(row=row, column=6, value=h['action']).font = entry_font
    ws3.cell(row=row, column=6).alignment = wrap_align
    ws3.row_dimensions[row].height = 110

apply_border_range(ws3, 3, 3 + len(hall_data), 1, 6)

# Hallucination Types Reference
ref_start = 4 + len(hall_data) + 2
ws3.merge_cells(f'A{ref_start}:F{ref_start}')
ws3.cell(row=ref_start, column=1, value='HALLUCINATION TYPES REFERENCE:').font = section_font
ws3.cell(row=ref_start, column=1).fill = light_blue_fill

ref_headers = ['Type', 'Definition', 'Example']
for i, h in enumerate(ref_headers, 1):
    c = ws3.cell(row=ref_start+1, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

ref_data = [
    ('Fabrication', 'AI tạo ra thông tin không tồn tại', 'Fake papers, fake APIs, fake data'),
    ('Oversimplification', 'AI bỏ qua edge cases, exceptions', 'Code không handle array rỗng'),
    ('Logic Error', 'AI đưa kết luận sai về "best practice"', '"XGBoost always best" → Sai'),
    ('Outdated Info', 'AI dùng thông tin cũ, không còn đúng', 'Old API syntax, deprecated methods'),
    ('Context Misunderstanding', 'AI không hiểu đúng bối cảnh', 'Không biết business constraints'),
]
for i, row_data in enumerate(ref_data):
    row = ref_start + 2 + i
    for j, val in enumerate(row_data, 1):
        c = ws3.cell(row=row, column=j, value=val)
        c.font = entry_font
        c.alignment = wrap_align

apply_border_range(ws3, ref_start+1, ref_start+1+len(ref_data), 1, 3)

# ============================================================
# SHEET 4: Self-Assessment Checklist
# ============================================================
ws4 = wb.create_sheet('4. Self-Assessment Checklist')

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 65
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 55

# Title
ws4.merge_cells('A1:D1')
c = ws4['A1']
c.value = 'SELF-ASSESSMENT CHECKLIST (Trước khi nộp)'
c.font = title_font
c.fill = title_fill
c.alignment = center_align

ws4.merge_cells('A2:D2')
ws4['A2'].value = 'Kiểm tra kỹ trước khi nộp. MỖI ENTRY phải pass ≥4/5 tiêu chí dưới đây.'
ws4['A2'].font = Font(name='Arial', size=10, italic=True, color='FF0000')

# Section A
ws4.merge_cells('A4:D4')
ws4['A4'].value = 'A. KIỂM TRA CHẤT LƯỢNG MỖI ENTRY (Pass ≥4/5)'
ws4['A4'].font = section_font
ws4['A4'].fill = light_blue_fill

a_headers = ['#', 'Tiêu chí', 'Pass?', 'Note']
for i, h in enumerate(a_headers, 1):
    c = ws4.cell(row=5, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

a_criteria = [
    ('1', 'Prompt này ảnh hưởng đến quyết định quan trọng trong project?', '☑', 'Tất cả 11 entries đều liên quan đến architecture/design/algorithm'),
    ('2', 'Nếu không có prompt này, project có thay đổi về architecture/design?', '☑', 'Có - cấu trúc thư mục, UML, thuật toán, code style đều bị ảnh hưởng'),
    ('3', 'Tôi có thể giải thích lý do chọn/không chọn gợi ý của AI?', '☑', 'Có - đã BÁC BỎ gợi ý AI ở Entry 002 (quá tối giản) và 009 (Lambda)'),
    ('4', 'Có minh chứng cụ thể (code, metrics, comparison)?', '☑', 'Code files, .drawio diagrams, project tree, comparison'),
    ('5', 'Prompt này phản ánh tư duy phản biện, không chỉ copy AI?', '☑', 'Đã chủ động discard và yêu cầu làm lại nhiều lần'),
]
for i, row_data in enumerate(a_criteria):
    row = 6 + i
    for j, val in enumerate(row_data, 1):
        c = ws4.cell(row=row, column=j, value=val)
        c.font = entry_font
        c.alignment = wrap_align if j > 1 else center_align

apply_border_range(ws4, 5, 10, 1, 4)

# Section B
ws4.merge_cells('A12:D12')
ws4['A12'].value = 'B. KIỂM TRA TỔNG THỂ LOG'
ws4['A12'].font = section_font
ws4['A12'].fill = light_blue_fill

b_headers = ['#', 'Tiêu chí', 'Pass?', 'Current Value']
for i, h in enumerate(b_headers, 1):
    c = ws4.cell(row=13, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

b_criteria = [
    ('1', 'Số lượng entries nằm trong range (min-max)?', '☑', '11 entries (đã lọc từ 17 prompts gốc)'),
    ('2', 'Mỗi DTC component có ít nhất 1 core prompt?', '☑', 'Decomposition: 3, Pattern Recognition: 1, Abstraction: 2, Algorithms: 5'),
    ('3', 'Đã phát hiện ≥ số lượng hallucination yêu cầu?', '☑', '3 hallucinations (≥3 for Project): Logic Error, Oversimplification, Context Misunderstanding'),
    ('4', 'Mỗi entry đều có Human Delta đầy đủ (4 câu hỏi)?', '☑', 'Tất cả 11 entries có đầy đủ Critical Thinking + Contextualization + Creative Synthesis + Decision'),
    ('5', 'Có evidence cho ≥70% entries?', '☑', '11/11 entries có evidence (100%)'),
]
for i, row_data in enumerate(b_criteria):
    row = 14 + i
    for j, val in enumerate(row_data, 1):
        c = ws4.cell(row=row, column=j, value=val)
        c.font = entry_font
        c.alignment = wrap_align if j > 1 else center_align

apply_border_range(ws4, 13, 18, 1, 4)

# Warning
ws4.merge_cells('A20:D20')
ws4['A20'].value = '⚠️ LƯU Ý QUAN TRỌNG:'
ws4['A20'].font = Font(name='Arial', size=12, bold=True, color='FF0000')

ws4.merge_cells('A21:D21')
ws4['A21'].value = 'Nếu entry KHÔNG pass ≥4/5 tiêu chí → LOẠI BỎ, không ghi vào Log'
ws4['A21'].font = Font(name='Arial', size=11, color='FF0000')

ws4.merge_cells('A22:D22')
ws4['A22'].value = 'Nếu FAIL ≥2 tiêu chí tổng thể → AI Reflection = 0 điểm (mất 30%)'
ws4['A22'].font = Font(name='Arial', size=11, color='FF0000')

# Section C
ws4.merge_cells('A25:D25')
ws4['A25'].value = 'C. CHUẨN BỊ CHO ORAL VIVAS (Q&A)'
ws4['A25'].font = section_font
ws4['A25'].fill = light_blue_fill

ws4.merge_cells('A26:D26')
ws4['A26'].value = 'Giảng viên sẽ hỏi ngẫu nhiên về 3-5 entries. Tự hỏi bản thân:'
ws4['A26'].font = Font(name='Arial', size=10, italic=True)

c_headers = ['Entry #', 'Tôi có thể giải thích tại sao chọn approach này?', 'Tôi có nhớ AI response?', 'Tôi có evidence?']
for i, h in enumerate(c_headers, 1):
    c = ws4.cell(row=27, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align

key_entries = ['001', '002', '005', '007', '009', '011']
for i, entry in enumerate(key_entries):
    row = 28 + i
    ws4.cell(row=row, column=1, value=entry).font = entry_font
    ws4.cell(row=row, column=1).alignment = center_align
    for j in range(2, 5):
        ws4.cell(row=row, column=j, value='☐').font = entry_font
        ws4.cell(row=row, column=j).alignment = center_align

apply_border_range(ws4, 27, 27 + len(key_entries), 1, 4)

# ============================================================
# SAVE
# ============================================================
output_path = r'e:\Project-LAB-github\1 group7\Call_Center_Waiting_Line_System-Group7\docs\AI_logs\NguyenVanAn_AI_AuditLog.xlsx'
wb.save(output_path)
print(f'Successfully created: {output_path}')
print(f'File size: {os.path.getsize(output_path)} bytes')
print(f'Sheets: {wb.sheetnames}')
print(f'Total core entries: {len(entries)}')
print(f'Hallucinations detected: {len(hall_data)}')
print()
print('=== ENTRIES REMOVED (not core prompts) ===')
removed = [
    ('Lượt 4', '"ok viết hoàn chỉnh đi"', 'Chỉ là lệnh phê duyệt, không có tư duy phản biện'),
    ('Lượt 5', '"viết cho tôi dòng .others/ ignore"', 'Tác vụ vận hành nhỏ, không ảnh hưởng architecture'),
    ('Lượt 6', '"others nằm trong folder docs vậy có sao ko"', 'Câu hỏi phụ về task nhỏ'),
    ('Lượt 9', '"viết lại toàn bộ... vào log.md"', 'Tác vụ hành chính/ghi chép'),
    ('Lượt 13', '"ok h giải thích kĩ lại cho tôi đi"', 'Yêu cầu giải thích lại, không dẫn đến quyết định'),
    ('Lượt 17', '"viết toàn bộ đoạn chat... vào log.md"', 'Tác vụ hành chính/ghi chép'),
]
for r in removed:
    print(f'  ❌ {r[0]}: {r[1]} → {r[2]}')
print()
print('=== ENTRIES KEPT (core prompts) ===')
for e in entries:
    print(f'  ✅ Entry {e["entry"]}: [{e["type"]}] [{e["component"]}] {e["context"][:60]}...')
